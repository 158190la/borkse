"""
API endpoint para Railway - LD Wealth Management
"""

from flask import Flask, jsonify, request
from flask_cors import CORS
import asyncio
import threading
import os
import requests as req

app = Flask(__name__)
CORS(app)

@app.after_request
def add_cors(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    return response

@app.route('/', defaults={'path': ''}, methods=['OPTIONS'])
@app.route('/<path:path>', methods=['OPTIONS'])
def options_handler(path=''):
    response = app.make_default_options_response()
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    return response

def ensure_credentials():
    creds_env = os.environ.get('GOOGLE_CREDENTIALS')
    if creds_env:
        with open('credentials.json', 'w') as f:
            f.write(creds_env)
        return True
    return os.path.exists('credentials.json')

def run_async(coro):
    result = {}
    def target():
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result['value'] = loop.run_until_complete(coro)
            result['ok'] = True
        except Exception as e:
            result['ok'] = False
            result['error'] = str(e)
        finally:
            loop.close()
    t = threading.Thread(target=target)
    t.start()
    t.join()
    return result

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'service': 'ldwm-api'})

@app.route('/scrape', methods=['GET', 'POST'])
def scrape():
    try:
        if not ensure_credentials():
            return jsonify({'ok': False, 'msg': 'No se encontro GOOGLE_CREDENTIALS.'}), 500
        from borkse import main as scrape_main
        result = run_async(scrape_main())
        if result.get('ok'):
            return jsonify({'ok': True, 'msg': 'Scraper completado.'})
        return jsonify({'ok': False, 'msg': result.get('error', 'Error desconocido')}), 500
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route('/api/bonds', methods=['GET'])
def api_bonds():
    try:
        if not ensure_credentials():
            return jsonify({'ok': False, 'msg': 'No se encontro GOOGLE_CREDENTIALS.'}), 500
        from dashboard import load_bonds, process_bonds, fetch_treasury_yields
        tsy = fetch_treasury_yields()
        raw = load_bonds()
        bonds = process_bonds(raw, tsy)
        return jsonify(bonds)
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route('/api/market', methods=['GET'])
def api_market():
    from urllib.parse import unquote
    tickers_param = request.args.get('tickers', '')
    if not tickers_param:
        return jsonify({'ok': False, 'msg': 'No tickers provided'}), 400
    tickers_param = unquote(tickers_param)
    tickers = [t.strip() for t in tickers_param.split(',') if t.strip()]
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0',
        'Accept': 'application/json',
    }
    results = {}
    for ticker in tickers:
        data = None
        urls = [
            f'https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?interval=5m&range=1d',
            f'https://query2.finance.yahoo.com/v8/finance/chart/{ticker}?interval=5m&range=1d',
            f'https://query1.finance.yahoo.com/v7/finance/quote?symbols={ticker}',
        ]
        for url in urls:
            try:
                r = req.get(url, headers=headers, timeout=12)
                if r.status_code != 200:
                    continue
                d = r.json()
                if 'chart' in d:
                    res = (d.get('chart') or {}).get('result') or []
                    if res:
                        meta   = res[0]['meta']
                        price  = meta.get('regularMarketPrice') or meta.get('previousClose')
                        prev   = meta.get('chartPreviousClose') or meta.get('previousClose') or price
                        chg    = ((price - prev) / prev * 100) if (prev and price) else 0
                        q0     = ((res[0].get('indicators') or {}).get('quote') or [{}])[0]
                        closes = [v for v in (q0.get('close') or []) if v is not None]
                        data   = {'price': round(float(price), 6), 'chg': round(float(chg), 4), 'closes': closes[-50:]}
                        break
                elif 'quoteResponse' in d:
                    qs = (d.get('quoteResponse') or {}).get('result') or []
                    if qs:
                        q = qs[0]
                        price = q.get('regularMarketPrice')
                        chg   = q.get('regularMarketChangePercent') or 0
                        data  = {'price': round(float(price), 6) if price else None, 'chg': round(float(chg), 4), 'closes': []}
                        break
            except Exception:
                continue
        results[ticker] = data
    return jsonify({'ok': True, 'data': results})

def fetch_market_snapshot():
    tickers = {
        'SP500': '^GSPC', 'NASDAQ': '^IXIC', 'DOW': '^DJI',
        'VIX': '^VIX', 'MERVAL': '^MERV',
        'BRENT': 'BZ=F', 'WTI': 'CL=F', 'GOLD': 'GC=F', 'NATGAS': 'NG=F',
        'DXY': 'DX-Y.NYB', 'UST2Y': '^IRX', 'UST10Y': '^TNX', 'BTC': 'BTC-USD',
    }
    headers = {'User-Agent': 'Mozilla/5.0'}
    results = {}
    for name, ticker in tickers.items():
        try:
            url = f'https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?interval=1d&range=2d'
            r = req.get(url, headers=headers, timeout=8)
            meta  = r.json()['chart']['result'][0]['meta']
            price = meta.get('regularMarketPrice') or meta.get('previousClose', 0)
            prev  = meta.get('previousClose') or price
            chg   = ((price - prev) / prev * 100) if prev else 0
            results[name] = {'price': round(price, 2), 'chg': round(chg, 2)}
        except Exception:
            results[name] = {'price': 'N/D', 'chg': 0}
    return results

@app.route('/api/parte', methods=['GET', 'POST'])
def api_parte():
    try:
        import anthropic, json
        from datetime import datetime
        fecha = datetime.now().strftime('%A %d de %B de %Y')
        mkt   = fetch_market_snapshot()
        def fmt(k):
            v = mkt.get(k, {})
            p = v.get('price', 'N/D')
            c = v.get('chg', 0)
            return f"{p} ({'+' if c > 0 else ''}{c}%)"
        market_context = f"""
DATOS DE MERCADO ({fecha}):
S&P 500: {fmt('SP500')} | Nasdaq: {fmt('NASDAQ')} | Dow: {fmt('DOW')} | VIX: {fmt('VIX')}
Merval: {fmt('MERVAL')} | Brent: {fmt('BRENT')} | WTI: {fmt('WTI')} | Oro: {fmt('GOLD')}
Gas: {fmt('NATGAS')} | DXY: {fmt('DXY')} | UST2Y: {fmt('UST2Y')} | UST10Y: {fmt('UST10Y')} | BTC: {fmt('BTC')}
"""
        prompt = f"""Sos el analista financiero senior de LD Wealth Management.
Genera el Parte Diario de hoy ({fecha}) usando SOLO estos datos reales:
{market_context}
Responde SOLO con JSON con estas 12 claves (array de 4-8 strings c/u):
claves, senales, flash, panorama, fed, fiscal, comercio, geo, usa, latam, argentina, ldwm
Sin texto extra, sin backticks."""
        client  = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY'))
        message = client.messages.create(
            model='claude-opus-4-5', max_tokens=4000,
            messages=[{'role': 'user', 'content': prompt}]
        )
        raw   = message.content[0].text.strip().replace('```json','').replace('```','').strip()
        parte = json.loads(raw)
        return jsonify({'ok': True, 'data': parte})
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)}), 500


# ── /api/fci/* ─────────────────────────────────────────────────────────────────
# Fuente: https://api.pub.cafci.org.ar/pb_get  (Excel publico, sin auth)
#
# Estructura real del Excel (confirmada via /api/fci/debug):
#   Fila 1: titulo "Camara Argentina de Fondos Comunes de Inversion"
#   Fila 2: direccion
#   Fila 3: descripcion "Reporte: Planilla Diaria"
#   Fila 4: HEADERS REALES (columna 0 = "Fondo", _col1 = "Clasificacion", etc.)
#   Fila 5+: datos
#
# Columnas relevantes (por posicion, 0-indexed):
#   0  = Fondo (nombre fondo - clase, ej: "FIMA Premium - Clase A")
#   1  = Clasificacion (tipo horizonte)
#   4  = Fecha
#   5  = Valor (mil cuotapartes) = VCP x 1000
#   9  = Variacion cuotaparte %  = rendimiento diario
#   14 = Patrimonio
#   20 = Codigo CAFCI            = "fondoId;claseId"
#   23 = Sociedad Gerente
#   25 = Codigo de Moneda
#   36 = Moneda Fondo (nombre)
# ──────────────────────────────────────────────────────────────────────────────

def _safe_float(v):
    try:
        return round(float(v), 4) if v not in (None, '', 'N/A') else None
    except Exception:
        return None

_CAFCI_XLS_CACHE = {'fondos': None, 'ts': 0}

# Mapeo de posicion de columna (0-indexed) a nombre semantico
# Basado en la fila de headers real (fila 4 del Excel)
_COL_IDX = {
    'nombre':      0,   # "Fondo" -> "FIMA Premium - Clase A"
    'tipo':        1,   # "Clasificacion"
    'fecha':       4,   # "Fecha"
    'vcp':         5,   # "Valor (mil cuotapartes)"
    'rend_dia':    9,   # "Variacion cuotaparte %"
    'patrimonio':  14,  # "Patrimonio"
    'cafci_code':  20,  # "Codigo CAFCI" -> "fondoId;claseId"
    'gerente':     23,  # "Sociedad Gerente"
    'cod_moneda':  25,  # "Codigo de Moneda"
    'moneda':      36,  # "Moneda Fondo"
}

def _cafci_load_fondos():
    """
    Descarga el Excel publico de CAFCI, salta las 3 filas de encabezado,
    usa la fila 4 como header y parsea los datos desde la fila 5.
    Cachea 4 horas.
    """
    import time, io
    import openpyxl
    now = time.time()
    if _CAFCI_XLS_CACHE['fondos'] and now - _CAFCI_XLS_CACHE['ts'] < 4 * 3600:
        return _CAFCI_XLS_CACHE['fondos']

    r = req.get(
        'https://api.pub.cafci.org.ar/pb_get', timeout=25,
        headers={'User-Agent': 'Mozilla/5.0', 'Referer': 'https://www.cafci.org.ar/'}
    )
    if r.status_code != 200:
        raise Exception(f'CAFCI planilla HTTP {r.status_code}')

    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    ws = wb.active

    fondos = {}
    header_row_found = False
    skip_count = 0

    for row in ws.iter_rows(values_only=True):
        # Las primeras filas no-vacias son: titulo, direccion, descripcion, headers
        # Detectamos la fila de headers buscando la celda que diga "Fondo" en col 0
        # y "Clasificacion" o similar en col 1
        if not header_row_found:
            if row[0] is not None and str(row[0]).strip().lower() == 'fondo':
                header_row_found = True
            # tambien saltamos si es la fila con el titulo real (col1 = Clasificacion)
            elif row[1] is not None and 'clasificaci' in str(row[1]).lower():
                header_row_found = True
            continue  # salta hasta encontrar el header

        # A partir de aca son filas de datos
        cols = list(row)
        nombre_raw = cols[_COL_IDX['nombre']] if len(cols) > _COL_IDX['nombre'] else None
        cafci_code = cols[_COL_IDX['cafci_code']] if len(cols) > _COL_IDX['cafci_code'] else None

        if not nombre_raw or not cafci_code:
            continue

        # Codigo CAFCI: puede ser "847;2409" o numero entero/float
        fid, cid = None, None
        code_str = str(cafci_code).strip()
        if ';' in code_str:
            parts = code_str.split(';')
            if len(parts) >= 2:
                try:
                    fid = int(float(parts[0]))
                    cid = int(float(parts[1]))
                except Exception:
                    pass
        else:
            # Algunos archivos tienen solo el ID del fondo (sin clase separada)
            # En ese caso usamos el codigo como fondoId y 0 como claseId placeholder
            try:
                fid = int(float(code_str))
                cid = 0
            except Exception:
                pass

        if fid is None:
            continue

        key = f'{fid};{cid}'
        vcp_raw = cols[_COL_IDX['vcp']] if len(cols) > _COL_IDX['vcp'] else None
        rend_raw = cols[_COL_IDX['rend_dia']] if len(cols) > _COL_IDX['rend_dia'] else None
        pat_raw  = cols[_COL_IDX['patrimonio']] if len(cols) > _COL_IDX['patrimonio'] else None
        ger_raw  = cols[_COL_IDX['gerente']] if len(cols) > _COL_IDX['gerente'] else None
        tipo_raw = cols[_COL_IDX['tipo']] if len(cols) > _COL_IDX['tipo'] else None
        mon_raw  = cols[_COL_IDX['moneda']] if len(cols) > _COL_IDX['moneda'] else None

        fondos[key] = {
            'fondoId':    fid,
            'claseId':    cid,
            'nombre':     str(nombre_raw).strip(),
            'gerente':    str(ger_raw).strip() if ger_raw else '',
            'tipo':       str(tipo_raw).strip() if tipo_raw else '',
            'moneda':     str(mon_raw).strip() if mon_raw else 'ARS',
            'vcp':        _safe_float(vcp_raw),
            'patrimonio': _safe_float(pat_raw),
            'rendimientos': {
                'day':   _safe_float(rend_raw),
                'week':  None,
                'month': None,
                'ytd':   None,
                'year':  None,
            },
        }

    wb.close()

    if not fondos:
        raise Exception('No se parseo ningun fondo del Excel CAFCI')

    _CAFCI_XLS_CACHE['fondos'] = fondos
    _CAFCI_XLS_CACHE['ts'] = now
    return fondos


@app.route('/api/fci/search', methods=['GET'])
def api_fci_search():
    q = request.args.get('q', '').strip().lower()
    if not q or len(q) < 2:
        return jsonify({'ok': False, 'msg': 'q requerido (min 2 chars)'}), 400
    try:
        fondos = _cafci_load_fondos()
        results = [f for f in fondos.values() if q in f['nombre'].lower()][:40]
        results.sort(key=lambda f: (0 if f['nombre'].lower().startswith(q) else 1, f['nombre']))
        return jsonify({'ok': True, 'data': results, 'total': len(fondos)})
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)}), 500


@app.route('/api/fci/ficha', methods=['GET'])
def api_fci_ficha():
    fid_s = request.args.get('fondo')
    cid_s = request.args.get('clase')
    if not fid_s or not cid_s:
        return jsonify({'ok': False, 'msg': 'fondo y clase requeridos'}), 400
    try:
        fid, cid = int(fid_s), int(cid_s)
        fondos = _cafci_load_fondos()
        key = f'{fid};{cid}'
        f = fondos.get(key)
        if not f:
            # Buscar solo por fondoId si claseId no matchea
            matches = [v for k, v in fondos.items() if v['fondoId'] == fid]
            if matches:
                f = matches[0]
        if not f:
            return jsonify({'ok': False, 'msg': f'Fondo {fid}/{cid} no encontrado'}), 404
        return jsonify({'ok': True, 'data': f})
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)}), 500


@app.route('/api/fci/historico', methods=['GET'])
def api_fci_historico():
    fid   = request.args.get('fondo')
    cid   = request.args.get('clase')
    from_d = request.args.get('desde')
    to_d   = request.args.get('hasta')
    if not all([fid, cid, from_d, to_d]):
        return jsonify({'ok': False, 'msg': 'fondo clase desde hasta requeridos'}), 400
    try:
        def fmt(d):
            y, m, day = d.split('-')
            return f'{day}-{m}-{y}'
        url = f'https://api.pub.cafci.org.ar/fondo/{fid}/clase/{cid}/rendimiento/{fmt(from_d)}/{fmt(to_d)}'
        r = req.get(url, timeout=20,
                    headers={'User-Agent': 'Mozilla/5.0', 'Referer': 'https://www.cafci.org.ar/'})
        if r.status_code != 200:
            return jsonify({'ok': True, 'data': []})
        text = r.text.strip()
        if not text or text[0] not in ('{', '['):
            return jsonify({'ok': True, 'data': []})
        items = r.json().get('data', [])
        series = [
            {'fecha': it['fecha'], 'vcp': float(it['vcp'])}
            for it in items
            if isinstance(it, dict) and it.get('fecha') and it.get('vcp') is not None
        ]
        return jsonify({'ok': True, 'data': series})
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)}), 500


@app.route('/api/fci/debug', methods=['GET'])
def api_fci_debug():
    """Muestra la estructura real del Excel y los primeros 3 fondos parseados."""
    import io
    import openpyxl
    try:
        r = req.get('https://api.pub.cafci.org.ar/pb_get', timeout=25,
                    headers={'User-Agent': 'Mozilla/5.0', 'Referer': 'https://www.cafci.org.ar/'})
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        ws = wb.active
        raw_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            raw_rows.append([str(c) if c is not None else None for c in row])
            if i >= 5:
                break
        wb.close()
        fondos = _cafci_load_fondos()
        sample = list(fondos.values())[:5]
        return jsonify({
            'ok': True,
            'total_fondos': len(fondos),
            'first_6_raw_rows': raw_rows,
            'sample_fondos': sample
        })
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)}), 500


@app.route('/api/treasury', methods=['GET'])
def api_treasury():
    try:
        from dashboard import fetch_treasury_yields
        return jsonify(fetch_treasury_yields())
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
